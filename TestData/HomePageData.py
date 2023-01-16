import openpyxl


book = openpyxl.load_workbook("/home/amjathhassan/Documents/PythonDemo.xlsx")
sheet = book.active


class HomePageData:

    test_HomePage_data = [{"first_name": "Rahul", "email": "rahul@mail.com", "password": "111111", "gender": "Male"},
                          {"first_name": "Jitha", "email": "jitha@mail.com", "password": "222222", "gender": "Female"}]

    @staticmethod
    def getData(test_case_name):

        Dict = {}
        List = []
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == test_case_name:
                for col in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
                    List.append(Dict)
        print(Dict)
        return [Dict]